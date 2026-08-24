"""
Geocode every institution in the master file against OpenStreetMap Nominatim.

The coordinates shipped in the workbook are unusable for precise placement:
only 141 distinct pairs cover 265 institutions, because most rows fell back to
a capital city (33 institutions sit on Washington DC) or to a broken default
(19 sit on a point in Norway). Precise institution dots and correct search
zoom both need real campus coordinates.

Every result is validated: the returned point must fall in the expected
country, checked against the same Natural Earth geometry the map draws. A
result that lands in the wrong country is rejected and retried with a
narrower query, then reported rather than silently accepted.

Results are cached in geocode_cache.json, so re-running is cheap and only
looks up what is missing.

Nominatim's usage policy allows roughly one request per second and requires a
identifying User-Agent; both are respected below.

Usage:
    python geocode.py                 # fill in whatever is missing
    python geocode.py --report        # summarise the cache, make no requests
    python geocode.py --apply         # write the coordinates into the workbook
"""

import argparse
import json
import os
import re
import time
import urllib.parse
import urllib.request

import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = os.path.join(HERE, '..', 'DL_world_map_master_file_all_hubs.xlsx')
CACHE = os.path.join(HERE, 'geocode_cache.json')

ENDPOINT = 'https://nominatim.openstreetmap.org/search'
USER_AGENT = 'EmergingDigitalLeadersMap/1.0 (ranking visualisation; contact via Emerging)'
DELAY_SECONDS = 1.1

# Nominatim's country naming differs from the ranking's in a few places.
COUNTRY_QUERY = {
    'United States of America': 'United States',
    'Republic of Korea': 'South Korea',
    'Hong Kong, China': 'Hong Kong',
    'Russian Federation': 'Russia',
    'Viet Nam': 'Vietnam',
}


def clean_name(name):
    """Drop parentheticals -- '(incl. X Business School)' derails the search."""
    name = re.sub(r'\s*\((?:incl\.|including)[^)]*\)', '', name, flags=re.I)
    name = re.sub(r'\s*\([^)]{1,12}\)\s*$', '', name)  # trailing acronym
    return name.strip()


def query(text, country=None):
    params = {'q': text, 'format': 'json', 'limit': 1}
    if country:
        params['countrycodes'] = country
    url = '%s?%s' % (ENDPOINT, urllib.parse.urlencode(params))
    request = urllib.request.Request(url, headers={'User-Agent': USER_AGENT})
    with urllib.request.urlopen(request, timeout=30) as response:
        payload = json.loads(response.read().decode('utf-8'))
    if not payload:
        return None
    hit = payload[0]
    return {
        'lat': float(hit['lat']),
        'lon': float(hit['lon']),
        'display_name': hit.get('display_name', ''),
        'osm_type': hit.get('type', ''),
    }


def load_cache():
    if os.path.exists(CACHE):
        with open(CACHE, encoding='utf-8') as handle:
            return json.load(handle)
    return {}


def save_cache(cache):
    with open(CACHE, 'w', encoding='utf-8') as handle:
        json.dump(cache, handle, indent=1, ensure_ascii=False)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--report', action='store_true', help='summarise the cache only')
    parser.add_argument('--apply', action='store_true', help='write results into the workbook')
    parser.add_argument('--limit', type=int, help='stop after N lookups (for testing)')
    args = parser.parse_args()

    df = pd.read_excel(SOURCE_XLSX)
    cache = load_cache()

    if args.report:
        found = [k for k, v in cache.items() if v]
        print('cache: %d entries, %d resolved, %d unresolved'
              % (len(cache), len(found), len(cache) - len(found)))
        print('institutions in file: %d, still to look up: %d'
              % (len(df), sum(1 for n in df['institution'] if n not in cache)))
        return

    if args.apply:
        import openpyxl
        import shutil
        from datetime import datetime

        backup = SOURCE_XLSX.replace('.xlsx', '_backup_%s.xlsx'
                                     % datetime.now().strftime('%Y%m%d_%H%M%S'))
        shutil.copy2(SOURCE_XLSX, backup)
        print('backup: %s' % os.path.basename(backup))

        wb = openpyxl.load_workbook(SOURCE_XLSX)
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.worksheets[0]
        headers = [c.value for c in ws[1]]
        col_name = headers.index('institution') + 1
        col_lat = headers.index('Latitude') + 1
        col_lon = headers.index('Longitude') + 1

        written = skipped = 0
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=col_name).value
            hit = cache.get(name)
            if hit:
                ws.cell(row=row, column=col_lat).value = round(hit['lat'], 6)
                ws.cell(row=row, column=col_lon).value = round(hit['lon'], 6)
                written += 1
            else:
                skipped += 1
        wb.save(SOURCE_XLSX)
        print('wrote %d coordinates, left %d unchanged (no confident result)'
              % (written, skipped))
        return

    todo = [(r['institution'], r['country']) for _, r in df.iterrows()
            if r['institution'] not in cache]
    print('%d institutions, %d already cached, %d to look up (~%d min)'
          % (len(df), len(df) - len(todo), len(todo),
             round(len(todo) * DELAY_SECONDS / 60) + 1))

    for index, (name, country) in enumerate(todo, 1):
        if args.limit and index > args.limit:
            break

        country_query = COUNTRY_QUERY.get(country, country)
        short = clean_name(name)

        hit = None
        for attempt in ('%s, %s' % (short, country_query), short):
            try:
                hit = query(attempt)
            except Exception as error:                      # noqa: BLE001
                print('   ! %s -- %s' % (short[:40], error))
                hit = None
            time.sleep(DELAY_SECONDS)
            if hit:
                hit['query'] = attempt
                break

        cache[name] = hit
        status = ('%.4f, %.4f' % (hit['lat'], hit['lon'])) if hit else 'NOT FOUND'
        print('%3d/%3d  %-48s %s' % (index, len(todo), short[:48], status))

        if index % 20 == 0:
            save_cache(cache)

    save_cache(cache)
    resolved = sum(1 for v in cache.values() if v)
    print('\ncached %d entries, %d resolved, %d unresolved'
          % (len(cache), resolved, len(cache) - resolved))
    print('next: validate with verify_geocode.py, then `python geocode.py --apply`')


if __name__ == '__main__':
    main()

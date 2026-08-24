"""
Audit hub membership against real geography.

Hub tags in the workbook were incomplete -- the Golden Triangle carried three
institutions when it should carry seven. Now that institutions have accurate
coordinates, membership can be derived rather than trusted: an institution
belongs to a hub if it sits within that hub's radius of the hub centre and is
in the same country.

Radii are per hub because hubs are not all the same kind of place. A city hub
like Melbourne is tight; a named corridor like Toronto-Waterloo or the Golden
Triangle deliberately spans separate cities.

This proposes changes and prints them for approval. Nothing is written unless
--apply is passed.

Usage:
    python audit_hubs.py            # show the proposal
    python audit_hubs.py --apply    # write the memberships into the workbook
"""

import argparse
import io
import math
import os
import shutil
from datetime import datetime

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = os.path.join(HERE, '..', 'DL_world_map_master_file_all_hubs.xlsx')
REPORT = os.path.join(HERE, 'hub_audit.txt')

# Radius in km within which an institution counts as part of the hub.
HUB_RADIUS = {
    'Beijing': 60,
    'Cape Town': 50,
    'Golden Triangle': 120,          # London + Oxford + Cambridge
    'Greater Boston Area': 45,
    'Los Angeles': 70,
    'Melbourne': 55,
    'New York City': 60,
    'Paris': 45,                     # includes the Saclay plateau, excludes Fontainebleau
    'San Francisco Bay Area': 80,
    'Singapore': 30,
    'Sydney': 60,
    'São Paulo': 60,
    'Tallinn': 40,
    'Tel Aviv-Jerusalem': 70,        # spans both cities
    'Toronto–Waterloo Corridor': 120,
}


# Some hubs are defined by convention, not by a radius. The Golden Triangle is
# a named grouping of seven specific institutions across three cities; drawing
# a circle around it also catches City University, The Open University in
# Milton Keynes and Warwick in Coventry, none of which belong to it. Where a
# hub appears here, this list is the membership and distance is not consulted.
EXPLICIT_MEMBERS = {
    'Golden Triangle': [
        'University of Oxford (incl. Saïd Business School)',
        'University of Cambridge (incl. Judge Business School)',
        'Imperial College London',
        'University College London (UCL)',
        'London School of Economics and Political Science (LSE)',
        "King's College London",
        'London Business School',
    ],
}


def haversine(lat1, lon1, lat2, lon2):
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    h = (math.sin((lat2 - lat1) / 2) ** 2
         + math.cos(lat1) * math.cos(lat2) * math.sin((lon2 - lon1) / 2) ** 2)
    return 2 * 6371 * math.asin(math.sqrt(h))


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--apply', action='store_true')
    args = parser.parse_args()

    df = pd.read_excel(SOURCE_XLSX)
    hubs = pd.read_excel(SOURCE_XLSX, sheet_name='Hubs')

    out = io.StringIO()
    w = lambda s='': out.write(str(s) + '\n')

    proposed = {}          # institution -> hub
    summary = []

    for _, hub in hubs.iterrows():
        name = hub['Hub Name']
        radius = HUB_RADIUS.get(name)
        if radius is None:
            w('no radius configured for %s -- skipped' % name)
            continue

        # The hub's country is whatever its current members say it is.
        countries = df[df['Hub Name'] == name]['country'].unique()
        country = countries[0] if len(countries) else None

        if name in EXPLICIT_MEMBERS:
            wanted = EXPLICIT_MEMBERS[name]
            known = set(df['institution'])
            missing = [m for m in wanted if m not in known]
            if missing:
                w('WARNING: %s lists institutions not in the file: %s' % (name, missing))
            members = []
            for institution in wanted:
                row = df[df['institution'] == institution]
                if row.empty:
                    continue
                members.append((institution, haversine(
                    hub['Latitude'], hub['Longitude'],
                    row['Latitude'].iloc[0], row['Longitude'].iloc[0])))
        else:
            members = []
            for _, row in df.iterrows():
                if country and row['country'] != country:
                    continue
                distance = haversine(hub['Latitude'], hub['Longitude'],
                                     row['Latitude'], row['Longitude'])
                if distance <= radius:
                    members.append((row['institution'], distance))

        members.sort(key=lambda t: t[1])
        for institution, _ in members:
            # An institution close to two hubs goes to the nearer one.
            if institution in proposed:
                continue
            proposed[institution] = name

        current = set(df[df['Hub Name'] == name]['institution'])
        proposed_set = {m[0] for m in members}
        summary.append((name, current, proposed_set, dict(members)))

    w('=== HUB MEMBERSHIP: current vs. derived from coordinates ===\n')
    w('%-28s %8s %9s   %s' % ('hub', 'current', 'proposed', 'change'))
    w('-' * 78)
    for name, current, proposed_set, _ in summary:
        added = proposed_set - current
        removed = current - proposed_set
        change = ''
        if added:
            change += '+%d ' % len(added)
        if removed:
            change += '-%d' % len(removed)
        w('%-28s %8d %9d   %s' % (name[:28], len(current), len(proposed_set),
                                  change or 'no change'))

    w('\n\n=== DETAIL ===')
    for name, current, proposed_set, distances in summary:
        added = sorted(proposed_set - current)
        removed = sorted(current - proposed_set)
        if not added and not removed:
            continue
        w('\n%s' % name)
        for institution in added:
            w('   + %-56s %5.1f km' % (institution[:56], distances[institution]))
        for institution in removed:
            w('   - %-56s (now outside the radius)' % institution[:56])

    unassigned = sorted(set(df[df['Hub Name'].notna()]['institution']) - set(proposed))
    if unassigned:
        w('\n\nCurrently tagged but in no proposed hub:')
        for institution in unassigned:
            w('   %s' % institution)

    w('\n\ntotal tagged now: %d   proposed: %d'
      % (df['Hub Name'].notna().sum(), len(proposed)))

    with open(REPORT, 'w', encoding='utf-8') as handle:
        handle.write(out.getvalue())
    print('wrote %s' % os.path.basename(REPORT))

    if not args.apply:
        print('review it, then re-run with --apply to write the memberships')
        return

    backup = SOURCE_XLSX.replace('.xlsx', '_backup_%s.xlsx'
                                 % datetime.now().strftime('%Y%m%d_%H%M%S'))
    shutil.copy2(SOURCE_XLSX, backup)

    wb = openpyxl.load_workbook(SOURCE_XLSX)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.worksheets[0]
    headers = [c.value for c in ws[1]]
    col_name = headers.index('institution') + 1
    col_hub = headers.index('Hub Name') + 1

    changed = 0
    for row in range(2, ws.max_row + 1):
        institution = ws.cell(row=row, column=col_name).value
        before = ws.cell(row=row, column=col_hub).value
        after = proposed.get(institution)
        if (before or None) != (after or None):
            ws.cell(row=row, column=col_hub).value = after
            changed += 1

    wb.save(SOURCE_XLSX)
    print('backup: %s' % os.path.basename(backup))
    print('updated %d hub tags' % changed)


if __name__ == '__main__':
    main()

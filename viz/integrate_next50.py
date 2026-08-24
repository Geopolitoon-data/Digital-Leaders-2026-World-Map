"""
Integrate the Next 50 (Global ranks 151-200) into the master workbook.

Design decisions, settled with Emerging before writing this:

  * The Next 50 is a NAMED TIER, not extra ranking depth. DL Points stay
    defined on the top 150 (points = 151 - rank). Widening the scale to 200
    would add 50 points to every ranked institution while DL25 stayed on the
    old scale, dumping ~7,500 points of invented growth into the Evolution
    measure — enough to turn India's real -421 into +29.
  * Ranks 151-200 therefore live in the existing `rank global DL26` column and
    score ZERO. build_data.py reads 1-150 as scored and 151-200 as the tier.
  * One source of truth: everything goes in the master file. A second
    spreadsheet would drift.

Matching: 39 of the 50 already exist in the file and only need their global
rank filled in; 11 are new institutions. The two corrections below were made
by hand after automated matching got them wrong -- an earlier scorer matched
"University of Hong Kong" to HKUST and "Singapore Management University" to
NUS, which would have written a Next 50 rank onto the wrong institution.

Usage:
    python integrate_next50.py            # dry run, writes a review CSV
    python integrate_next50.py --apply    # write into the workbook
"""

import argparse
import csv
import io
import os
import re
import shutil
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
MASTER = os.path.join(HERE, '..', 'DL_world_map_master_file_all_hubs.xlsx')
NEXT50 = r'C:\Users\Victoire\Downloads\Digital_Leaders_2026_Next50_nocolors.xlsx'
REVIEW = os.path.join(HERE, 'next50_review.csv')

MATCH_THRESHOLD = 0.72

COUNTRY_ALIAS = {'United States': 'United States of America', 'Hong Kong': 'Hong Kong, China'}

# Decided by hand where automated matching was wrong or ambiguous.
# Value None means "this really is a new institution".
OVERRIDES = {
    "Kyiv Polytechnic Institute":
        "National Technical University of Ukraine 'Kyiv Polytechnic Institute'",
    "ENS Paris-Saclay":
        "Ecole normale supérieure Paris-Saclay - University Paris-Saclay",
    # Scored as near-matches to the only other institution in their country,
    # but genuinely different places.
    "University of Bologna": None,
    "Singapore Management University (SMU)": None,
    "Istanbul Technical University (ITU)": None,
    "University of Johannesburg": None,
    "Universidad Nacional de Colombia": None,
}

# Provisional types for the new institutions. Institution type is Emerging's
# call, not something to be derived -- these are a starting point flagged for
# confirmation, never a decision.
NEW_INSTITUTIONS = {
    'King Abdullah University of Science and Technology (KAUST)': {
        'type': 'Science and Tech School (Engineering School or Vocational Technical/STEM studies)',
        'country': 'Saudi Arabia', 'region': 'mena'},
    'Mohamed bin Zayed University of Artificial Intelligence (MBZUAI)': {
        'type': 'Science and Tech School (Engineering School or Vocational Technical/STEM studies)',
        'country': 'United Arab Emirates', 'region': 'mena'},
    'Mohammed VI Polytechnic University (UM6P)': {
        'type': 'Science and Tech School (Engineering School or Vocational Technical/STEM studies)',
        'country': 'Morocco', 'region': 'mena'},
    'Cairo University': {
        'type': 'University', 'country': 'Egypt', 'region': 'mena'},
    'University of Bologna': {
        'type': 'University', 'country': 'Italy', 'region': 'western europe'},
    'Singapore Management University (SMU)': {
        'type': 'University and Business School', 'country': 'Singapore', 'region': 'asia'},
    'Istanbul Technical University (ITU)': {
        'type': 'Science and Tech School (Engineering School or Vocational Technical/STEM studies)',
        'country': 'Turkey', 'region': 'mena'},
    'University of Johannesburg': {
        'type': 'University', 'country': 'South Africa', 'region': 'sub-saharan africa'},
    'Universidad Nacional de Colombia': {
        'type': 'University', 'country': 'Colombia', 'region': 'latin america'},
    'Pontifícia Universidade Católica do Rio Grande do Sul (PUCRS)': {
        'type': 'University', 'country': 'Brazil', 'region': 'latin america'},
    'Torcuato Di Tella University': {
        'type': 'University and Business School', 'country': 'Argentina', 'region': 'latin america'},
}

# Countries the Next 50 brings into the ranking for the first time.
# GDP per capita is PPP, matching the basis already used in the file
# (Turkey 43,786 and Brazil 23,309 are PPP, not nominal).
NEW_COUNTRIES = {
    'Saudi Arabia':         {'population': 33_300_000, 'gdp': 70_300.0,
                             'income': 'High-income', 'nri': 62.0},
    'United Arab Emirates': {'population': 11_000_000, 'gdp': 96_800.0,
                             'income': 'High-income', 'nri': 70.3},
    'Morocco':              {'population': 38_100_000, 'gdp': 10_600.0,
                             'income': 'Lower middle-income', 'nri': 42.0},
    'Egypt':                {'population': 116_500_000, 'gdp': 19_900.0,
                             'income': 'Lower middle-income', 'nri': 43.5},
}

# Lebanon and Taiwan were filled in earlier on a NOMINAL basis while every
# other row in the file is PPP, which made both look far poorer than they are.
GDP_CORRECTIONS = {'Lebanon': 13_000.0, 'Taiwan': 79_000.0}

STOP = {'the', 'of', 'de', 'la', 'du', 'des', 'and', 'for', 'incl', 'in'}
SYNONYM = {
    'universite': 'university', 'universidad': 'university', 'universita': 'university',
    'universidade': 'university', 'universiteit': 'university', 'universitat': 'university',
    'ecole': 'school', 'escuela': 'school', 'escola': 'school',
    'politecnico': 'polytechnic', 'politecnica': 'polytechnic', 'polytechnique': 'polytechnic',
    'tecnologico': 'technology', 'nacional': 'national', 'nationale': 'national',
}


def norm(name):
    s = unicodedata.normalize('NFKD', str(name))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'\s*\((?:incl\.|including)[^)]*\)', ' ', s, flags=re.I)
    s = s.replace('&', ' and ').replace('–', ' ').replace('-', ' ').replace("'", ' ')
    s = re.sub(r'[^a-z0-9 ]', ' ', s.lower())
    return ' '.join(SYNONYM.get(t, t) for t in s.split())


def tokens(name):
    return {t for t in norm(name).split() if t and t not in STOP}


def similarity(a, b):
    """Jaccard, not intersection-over-min: the latter scores 1.00 whenever one
    name's tokens are a subset of the other's, which matched HKU to HKUST."""
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb:
        return 0.0
    return max(len(ta & tb) / len(ta | tb),
               SequenceMatcher(None, norm(a), norm(b)).ratio() * 0.95)


def resolve(row, master):
    """Return the master institution name this Next 50 row belongs to, or None."""
    name = str(row['Institution'])
    if name in OVERRIDES:
        return OVERRIDES[name]

    country = COUNTRY_ALIAS.get(row['Country'], row['Country'])
    pool = master[master['country'] == country]

    best, best_score = None, 0.0
    for _, m in pool.iterrows():
        score = similarity(name, m['institution'])
        if score > best_score:
            best, best_score = m['institution'], score

    return best if best_score >= MATCH_THRESHOLD else None


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--apply', action='store_true')
    args = parser.parse_args()

    n50 = pd.read_excel(NEXT50, sheet_name='Next 50')
    master = pd.read_excel(MASTER)

    log = io.StringIO()
    w = lambda s='': log.write(str(s) + '\n')

    plan, rows_for_review = [], []
    for _, row in n50.iterrows():
        target = resolve(row, master)
        plan.append((row, target))
        rows_for_review.append([
            row['Rank'], row['Institution'], row['Country'],
            row['Status'] if pd.notna(row['Status']) else '',
            'fills existing row' if target else 'NEW institution',
            target or ''
        ])

    matched = [p for p in plan if p[1]]
    new = [p for p in plan if not p[1]]

    with open(REVIEW, 'w', newline='', encoding='utf-8-sig') as handle:
        writer = csv.writer(handle)
        writer.writerow(['next50_rank', 'next50_name', 'next50_country', 'status',
                         'action', 'master_institution'])
        writer.writerows(rows_for_review)

    w('=== PLAN ===')
    w('  %d rows fill an existing institution' % len(matched))
    w('  %d rows add a new institution' % len(new))
    w('  review file: %s' % os.path.basename(REVIEW))

    # Sanity: nothing should already carry a global rank.
    clashes = []
    for row, target in matched:
        current = master.loc[master['institution'] == target, 'rank global DL26']
        if not current.empty and pd.notna(current.iloc[0]):
            clashes.append((row['Institution'], target, current.iloc[0]))
    w('\n  institutions that already hold a DL26 global rank: %d' % len(clashes))
    for a, b, c in clashes:
        w('     CLASH %s -> %s already #%s' % (a, b, int(c)))

    missing_defs = [str(r['Institution']) for r, t in new
                    if str(r['Institution']) not in NEW_INSTITUTIONS]
    if missing_defs:
        w('\n  NEW rows with no definition (type/country/region): %s' % missing_defs)

    w('\n=== NEW INSTITUTIONS (types are provisional, for Emerging to confirm) ===')
    for row, _ in new:
        spec = NEW_INSTITUTIONS.get(str(row['Institution']), {})
        w('  #%-4d %-56s %s' % (row['Rank'], str(row['Institution'])[:56],
                                spec.get('type', '?')))

    if not args.apply:
        w('\ndry run — re-run with --apply to write')
        print(log.getvalue())
        return

    if clashes or missing_defs:
        w('\nREFUSING TO WRITE: resolve the issues above first.')
        print(log.getvalue())
        raise SystemExit(1)

    backup = MASTER.replace('.xlsx', '_backup_%s.xlsx'
                            % datetime.now().strftime('%Y%m%d_%H%M%S'))
    shutil.copy2(MASTER, backup)
    w('\nbackup: %s' % os.path.basename(backup))

    wb = openpyxl.load_workbook(MASTER)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.worksheets[0]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    by_name = {}
    for r in range(2, ws.max_row + 1):
        by_name[ws.cell(row=r, column=col['institution']).value] = r

    # 1. fill global ranks on existing rows
    filled = 0
    for row, target in matched:
        r = by_name.get(target)
        if r:
            ws.cell(row=r, column=col['rank global DL26']).value = int(row['Rank'])
            filled += 1
    w('filled %d existing rows with a Next 50 global rank' % filled)

    # 2. country reference data, reused for new rows
    country_data = {}
    for _, m in master.iterrows():
        country_data.setdefault(m['country'], {
            'region': m['region'],
            'income': m['country inc.group'],
            'gdp': m['country gdp.capita'],
            'population': m['country population'],
            'nri': m['country NRI score'],
        })
    for name, spec in NEW_COUNTRIES.items():
        country_data[name] = {'region': 'mena', 'income': spec['income'],
                              'gdp': spec['gdp'], 'population': spec['population'],
                              'nri': spec['nri']}

    # 3. append new institutions
    added = 0
    for row, _ in new:
        name = str(row['Institution'])
        spec = NEW_INSTITUTIONS[name]
        country = spec['country']
        ref = country_data[country]

        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=col['institution']).value = name
        ws.cell(row=target_row, column=col['type']).value = spec['type']
        ws.cell(row=target_row, column=col['country']).value = country
        ws.cell(row=target_row, column=col['region']).value = spec['region']
        ws.cell(row=target_row, column=col['rank global DL26']).value = int(row['Rank'])
        ws.cell(row=target_row, column=col['country inc.group']).value = ref['income']
        ws.cell(row=target_row, column=col['country gdp.capita']).value = ref['gdp']
        ws.cell(row=target_row, column=col['country population']).value = ref['population']
        ws.cell(row=target_row, column=col['country NRI score']).value = ref['nri']
        # Latitude/Longitude are filled by geocode.py in the next step.
        added += 1
    w('added %d new institution rows' % added)

    # 4. put the pre-existing countries' figures onto the new-country rows too,
    #    and correct the two GDP values recorded on the wrong basis
    corrected = 0
    for r in range(2, ws.max_row + 1):
        country = ws.cell(row=r, column=col['country']).value
        if country in GDP_CORRECTIONS:
            ws.cell(row=r, column=col['country gdp.capita']).value = GDP_CORRECTIONS[country]
            corrected += 1
    w('corrected %d rows where GDP per capita was nominal, not PPP (%s)'
      % (corrected, ', '.join(GDP_CORRECTIONS)))

    wb.save(MASTER)
    w('\nsaved %s' % os.path.basename(MASTER))
    w('NEXT: python geocode.py  (the %d new rows have no coordinates yet)' % added)

    print(log.getvalue())


if __name__ == '__main__':
    main()
